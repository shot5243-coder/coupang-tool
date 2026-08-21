# -*- coding: utf-8 -*-
"""
네이버 API HUB - 검색어트렌드(Search Trend) 조회 클라이언트
- 키워드별 검색량 추이(상대지수, 0~100)를 조회
- 실행하면 Client ID/Secret을 직접 입력받으므로 파일을 수정할 필요 없음
- 문서: NAVER API HUB > Data Lab > 검색어트렌드

주의: 이 스크립트는 인터넷 연결이 되는 사용자 PC에서 실행해야 합니다.
"""

import os
import json
import datetime
import requests

BASE_URL = "https://naverapihub.apigw.ntruss.com/search-trend/v1/search"

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_KEY_FILE = os.path.join(_SCRIPT_DIR, "naver_trend_key.txt")


def load_saved_keys():
    """이전에 저장해둔 (key_id, key) 튜플이 있으면 불러옴"""
    if os.path.exists(_KEY_FILE):
        with open(_KEY_FILE, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
            if len(lines) >= 2:
                return lines[0], lines[1]
    return None


def save_keys(key_id: str, key: str):
    with open(_KEY_FILE, "w", encoding="utf-8") as f:
        f.write(key_id.strip() + "\n")
        f.write(key.strip() + "\n")


def search_trend(key_id: str, key: str, keywords: list, start_date: str, end_date: str,
                  time_unit: str = "date", ages: list = None, gender: str = None,
                  timeout: int = 10) -> dict:
    """
    keywords: 최대 20개 (검색어 자체가 아니라 '키워드 그룹의 대표어'를 넣으면 됨)
    각 키워드는 자동으로 별도 그룹으로 나눠서 요청 (최대 5개 그룹까지 한 번에 비교 가능)
    time_unit: 'date'(일간) / 'week'(주간) / 'month'(월간)
    ages: 연령대 코드 리스트 (5살 단위). 1=0-12, 2=13-18, 3=19-24, 4=25-29, 5=30-34,
          6=35-39, 7=40-44, 8=45-49, 9=50-54, 10=55-59, 11=60세 이상
    gender: 'm'(남성) / 'f'(여성) / None(전체)
    """
    if len(keywords) > 5:
        raise ValueError("한 번에 비교 가능한 키워드는 최대 5개입니다.")

    body = {
        "startDate": start_date,
        "endDate": end_date,
        "timeUnit": time_unit,
        "keywordGroups": [
            {"groupName": kw, "keywords": [kw]} for kw in keywords
        ],
    }
    if ages:
        body["ages"] = ages
    if gender:
        body["gender"] = gender
    headers = {
        "X-NCP-APIGW-API-KEY-ID": key_id,
        "X-NCP-APIGW-API-KEY": key,
        "Content-Type": "application/json",
    }
    resp = requests.post(BASE_URL, headers=headers, data=json.dumps(body), timeout=timeout)
    if resp.status_code != 200:
        raise RuntimeError(f"API 오류 (HTTP {resp.status_code}): {resp.text}")
    return resp.json()


def summarize(result: dict) -> list:
    """
    결과에서 키워드별 최신 지수, 평균 지수, 추세(상승/하락/보합)를 뽑아 요약
    반환: [{"키워드":.., "최신지수":.., "평균지수":.., "추세":..}, ...]
    """
    summary = []
    for group in result.get("results", []):
        title = group.get("title", "")
        data = group.get("data", [])
        if not data:
            summary.append({"키워드": title, "최신지수": "-", "평균지수": "-", "추세": "데이터 없음"})
            continue
        ratios = [d["ratio"] for d in data]
        latest = ratios[-1]
        avg = sum(ratios) / len(ratios)
        # 최근 구간(뒤 25%) 평균과 이전 구간(앞 25%) 평균을 비교해 추세 판정
        n = len(ratios)
        head = ratios[: max(1, n // 4)]
        tail = ratios[-max(1, n // 4):]
        head_avg, tail_avg = sum(head) / len(head), sum(tail) / len(tail)
        if tail_avg > head_avg * 1.15:
            trend = "상승"
        elif tail_avg < head_avg * 0.85:
            trend = "하락"
        else:
            trend = "보합"
        summary.append({
            "키워드": title,
            "최신지수": round(latest, 1),
            "평균지수": round(avg, 1),
            "추세": trend,
        })
    return summary


def print_summary(summary: list, time_unit: str):
    print(f"\n(지수는 조회 기간 내 최댓값을 100으로 둔 상대값입니다 / 기준: {time_unit})")
    print("-" * 60)
    print(f"{'키워드':<20}{'최신지수':>10}{'평균지수':>10}{'추세':>10}")
    print("-" * 60)
    for row in summary:
        print(f"{row['키워드']:<20}{str(row['최신지수']):>10}{str(row['평균지수']):>10}{row['추세']:>10}")
    print("-" * 60)


def export_to_excel(result: dict, summary: list, out_path: str = "네이버_검색트렌드.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "요약"

    headers = ["키워드", "최신지수", "평균지수", "추세"]
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
    for r, row in enumerate(summary, start=2):
        for col, key in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=row[key])
            c.font, c.alignment, c.border = Font(name=FONT), center, border
    for i, w in enumerate([20, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 일자별 추이 시트 + 라인 차트
    ws2 = wb.create_sheet("일자별 추이")
    groups = result.get("results", [])
    if groups:
        periods = [d["period"] for d in groups[0].get("data", [])]
        ws2.cell(row=1, column=1, value="기간")
        for col, g in enumerate(groups, start=2):
            ws2.cell(row=1, column=col, value=g.get("title", ""))
        for r, period in enumerate(periods, start=2):
            ws2.cell(row=r, column=1, value=period)
            for col, g in enumerate(groups, start=2):
                data = g.get("data", [])
                ratio = data[r - 2]["ratio"] if r - 2 < len(data) else None
                ws2.cell(row=r, column=col, value=ratio)

        chart = LineChart()
        chart.title = "키워드별 검색 트렌드"
        chart.y_axis.title = "상대지수"
        chart.x_axis.title = "기간"
        max_row = len(periods) + 1
        max_col = len(groups) + 1
        data_ref = Reference(ws2, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
        cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 24, 12
        ws2.add_chart(chart, f"{get_column_letter(max_col + 2)}2")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    saved = load_saved_keys()
    if saved:
        KEY_ID, KEY = saved
        print(f"저장된 API 키({KEY_ID[:6]}...)를 자동으로 사용합니다.")
    else:
        KEY_ID = input("Client ID (X-NCP-APIGW-API-KEY-ID) 입력: ").strip()
        KEY = input("Client Secret (X-NCP-APIGW-API-KEY) 입력: ").strip()
        if KEY_ID and KEY:
            save = input("다음에도 자동으로 쓰도록 저장할까요? (y/n): ").strip().lower()
            if save == "y":
                save_keys(KEY_ID, KEY)
                print(f"저장 완료: {_KEY_FILE}")

    if not KEY_ID or not KEY:
        print("Client ID/Secret이 입력되지 않았습니다.")
    else:
        raw = input("비교할 키워드 입력 (쉼표로 구분, 최대 5개, 예: 캠핑의자, 캠핑테이블): ").strip()
        keywords = [k.strip() for k in raw.split(",") if k.strip()][:5]

        today = datetime.date.today()
        default_end = today - datetime.timedelta(days=1)
        default_start = default_end - datetime.timedelta(days=365)
        start_input = input(f"조회 시작일 (엔터 = {default_start}): ").strip() or str(default_start)
        end_input = input(f"조회 종료일 (엔터 = {default_end}): ").strip() or str(default_end)
        unit_input = input("조회 단위 date/week/month (엔터 = date): ").strip() or "date"

        if not keywords:
            print("키워드가 입력되지 않았습니다.")
        else:
            try:
                result = search_trend(KEY_ID, KEY, keywords, start_input, end_input, unit_input)
                summary = summarize(result)
                print_summary(summary, unit_input)
                print("\n엑셀로 저장 중...")
                out_path = export_to_excel(result, summary)
                print(f"저장 완료: {out_path} (일자별 추이 + 그래프 포함)")
            except Exception as e:
                print(f"오류 발생: {e}")

    # 더블클릭으로 실행했을 때 창이 바로 닫히지 않도록 대기
    input("\n종료하려면 Enter를 누르세요...")
