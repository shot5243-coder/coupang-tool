# -*- coding: utf-8 -*-
"""
검색어트렌드 기반 연령대별 관심도 조회기
- 쇼핑인사이트(카테고리 코드 필요) 대신, 이미 정상 작동 중인 검색어트렌드 API에
  연령대 필터를 걸어서 같은 정보를 얻는다. 카테고리 코드가 필요 없다.
- search_trend_client.py와 같은 폴더에 있어야 하며 API 키를 공유한다.

연령대 구간 (네이버 표준, 5살 단위 코드를 10살 단위로 묶음):
  10대: 13~18세 / 20대: 19~29세 / 30대: 30~39세 / 40대: 40~49세 / 50대: 50~59세 / 60대 이상: 60세~
"""

import datetime
import search_trend_client as trend

AGE_BUCKETS = {
    "10대": ["2"],
    "20대": ["3", "4"],
    "30대": ["5", "6"],
    "40대": ["7", "8"],
    "50대": ["9", "10"],
    "60대 이상": ["11"],
}


def age_trend(key_id: str, key: str, keyword: str, start_date: str, end_date: str,
              time_unit: str = "month") -> list:
    """
    키워드 하나에 대해 연령대별로 각각 조회해서 평균 지수를 비교.
    (각 호출은 그 연령대 구간 내에서 상대지수를 다시 산정하므로, 절대 비교보다는
    '어느 연령대에서 상대적으로 더 꾸준히 검색되는지' 참고용으로 활용)
    """
    summary = []
    for bucket_name, age_codes in AGE_BUCKETS.items():
        try:
            result = trend.search_trend(key_id, key, [keyword], start_date, end_date,
                                         time_unit, ages=age_codes)
            groups = result.get("results", [])
            if not groups or not groups[0].get("data"):
                summary.append({"연령대": bucket_name, "최신지수": 0.0, "평균지수": 0.0})
                continue
            ratios = [d["ratio"] for d in groups[0]["data"]]
            summary.append({
                "연령대": bucket_name,
                "최신지수": round(ratios[-1], 1),
                "평균지수": round(sum(ratios) / len(ratios), 1),
            })
        except Exception as e:
            summary.append({"연령대": bucket_name, "최신지수": "오류", "평균지수": str(e)})

    total = sum(r["평균지수"] for r in summary if isinstance(r["평균지수"], (int, float))) or 1
    for r in summary:
        if isinstance(r["평균지수"], (int, float)):
            r["비중(%)"] = round(r["평균지수"] / total * 100, 1)
        else:
            r["비중(%)"] = "-"

    summary.sort(key=lambda x: x["평균지수"] if isinstance(x["평균지수"], (int, float)) else -1, reverse=True)
    return summary


def print_summary(keyword: str, summary: list):
    print(f"\n[{keyword}] 연령대별 관심도 (비중 높은 순)")
    print("-" * 50)
    print(f"{'연령대':<12}{'최신지수':>10}{'평균지수':>10}{'비중(%)':>12}")
    print("-" * 50)
    for row in summary:
        print(f"{row['연령대']:<12}{str(row['최신지수']):>10}{str(row['평균지수']):>10}{str(row['비중(%)']):>12}")
    print("-" * 50)
    valid = [r for r in summary if isinstance(r["평균지수"], (int, float))]
    if valid:
        print(f"  ※ 가장 관심도 높은 연령대: {valid[0]['연령대']}")


def export_to_excel(keyword: str, summary: list, out_path: str = None):
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    out_path = out_path or os.path.join(trend._SCRIPT_DIR if hasattr(trend, "_SCRIPT_DIR")
                                         else ".", f"연령대별_{keyword}.xlsx")
    FONT = "Arial"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "연령별 관심도"

    headers = ["연령대", "최신지수", "평균지수", "비중(%)"]
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
    for r, row in enumerate(summary, start=2):
        for col, key_name in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col, value=row[key_name])
            c.font, c.alignment, c.border = Font(name=FONT), center, border
    for i, w in enumerate([14, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    chart = BarChart()
    chart.title = f"{keyword} - 연령대별 관심도"
    chart.y_axis.title = "평균지수"
    last_row = len(summary) + 1
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width, chart.height = 16, 10
    ws.add_chart(chart, "F2")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    saved = trend.load_saved_keys()
    if saved:
        KEY_ID, KEY = saved
        print(f"저장된 API 키({KEY_ID[:6]}...)를 자동으로 사용합니다.")
    else:
        KEY_ID = input("Client ID (X-NCP-APIGW-API-KEY-ID) 입력: ").strip()
        KEY = input("Client Secret (X-NCP-APIGW-API-KEY) 입력: ").strip()
        if KEY_ID and KEY:
            if input("저장할까요? (y/n): ").strip().lower() == "y":
                trend.save_keys(KEY_ID, KEY)

    if not KEY_ID or not KEY:
        print("Client ID/Secret이 입력되지 않았습니다.")
    else:
        keyword = input("분석할 키워드 입력 (1개, 예: 캠핑랜턴): ").strip()
        today = datetime.date.today()
        default_end = today - datetime.timedelta(days=1)
        default_start = default_end - datetime.timedelta(days=365)
        start_input = input(f"조회 시작일 (엔터 = {default_start}): ").strip() or str(default_start)
        end_input = input(f"조회 종료일 (엔터 = {default_end}): ").strip() or str(default_end)

        if not keyword:
            print("키워드가 입력되지 않았습니다.")
        else:
            try:
                summary = age_trend(KEY_ID, KEY, keyword, start_input, end_input)
                print_summary(keyword, summary)
                out_path = export_to_excel(keyword, summary)
                print(f"\n저장 완료: {out_path}")
            except Exception as e:
                print(f"오류 발생: {e}")

    input("\n종료하려면 Enter를 누르세요...")
