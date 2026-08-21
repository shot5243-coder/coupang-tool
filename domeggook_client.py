# -*- coding: utf-8 -*-
"""
도매꾹/도매매 OpenAPI 연동 클라이언트
- getItemList(상품리스트) API로 키워드 검색 → 상품명/가격/배송비/URL 자동 수집
- 실행하면 API_KEY를 직접 입력받으므로 파일을 수정할 필요 없음
- 문서: https://openapi.domeggook.com

주의: 이 스크립트는 인터넷 연결이 되는 사용자 PC에서 실행해야 합니다.
"""

import io
import time
import requests
from dataclasses import dataclass
from typing import List, Optional

import os

BASE_URL = "https://domeggook.com/ssl/api/"

DEFAULT_TARGET_MARGIN = 0.4        # 목표 마진율 (도매꾹은 40% 달성이 어려운 경우가 많으니 조정해서 쓰세요)
DEFAULT_COUPANG_FEE_RATE = 0.108   # 쿠팡 카테고리 수수료율 (기본 10.8%, 카테고리별로 4~10.9% 편차)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_KEY_FILE = os.path.join(_SCRIPT_DIR, "domeggook_key.txt")


def load_saved_key() -> Optional[str]:
    """이전에 저장해둔 API_KEY가 있으면 불러옴"""
    if os.path.exists(_KEY_FILE):
        with open(_KEY_FILE, "r", encoding="utf-8") as f:
            key = f.read().strip()
            return key or None
    return None


def save_key(api_key: str):
    """API_KEY를 로컬 파일에 저장 (다음 실행부터 자동으로 불러옴)"""
    with open(_KEY_FILE, "w", encoding="utf-8") as f:
        f.write(api_key.strip())


@dataclass
class DomeItem:
    no: str            # 상품번호
    title: str          # 상품명
    price: int          # 가격(원)
    unit_qty: int        # 최소구매수량
    seller_id: str       # 판매자 아이디
    deli_fee: int         # 배송비
    deli_who: str         # 배송비 부담주체 (P=구매자 등)
    from_oversea: bool    # 해외배송 여부
    url: str              # 상품 상세 URL
    thumb: str             # 썸네일 이미지 URL


class DomeggookClient:
    def __init__(self, api_key: str, market: str = "dome"):
        """
        api_key: 도매꾹에서 발급받은 API Key (aid 파라미터)
        market: 'dome'(도매꾹) 또는 'supply'(도매매)
        """
        self.api_key = api_key
        self.market = market

    def search_items(self, keyword: str, size: int = 50, page: int = 1,
                      sort: str = "rd", timeout: int = 10) -> List[DomeItem]:
        """
        키워드로 상품 목록 검색.
        sort: rd(등록일순) / sr(판매순-일부버전) 등, 문서 기준 기본값 rd 사용
        """
        params = {
            "ver": "4.1",
            "mode": "getItemList",
            "aid": self.api_key,
            "market": self.market,
            "om": "json",
            "kw": keyword,
            "sz": size,
            "pg": page,
            "so": sort,
        }
        resp = requests.get(BASE_URL, params=params, timeout=timeout)
        resp.raise_for_status()
        data = resp.json()

        root = data.get("domeggook", data)  # 응답 루트가 domeggook 키로 감싸져 있을 수 있음

        # 에러 응답 체크 (도매꾹은 실패 시 errors 노드를 반환)
        if "errors" in root:
            err = root["errors"]
            # "마지막 페이지 초과"처럼 결과가 그냥 없어서 나는 에러는 예외가 아니라 빈 목록으로 처리
            msg = str(err.get("message", "")) if isinstance(err, dict) else str(err)
            if "페이지" in msg or "목록이 없습니다" in msg:
                return []
            raise RuntimeError(f"도매꾹 API 오류: {err}")

        item_list = root.get("list", {}).get("item", [])
        if isinstance(item_list, dict):  # 결과가 1건이면 dict로 옴 → list로 통일
            item_list = [item_list]

        items = []
        for it in item_list:
            deli = it.get("deli", {}) or {}
            items.append(DomeItem(
                no=str(it.get("no", "")),
                title=(it.get("title") or "").strip(),
                price=int(it.get("price", 0) or 0),
                unit_qty=int(it.get("unitQty", 1) or 1),
                seller_id=it.get("id", ""),
                deli_fee=int(deli.get("fee", 0) or 0),
                deli_who=deli.get("who", ""),
                from_oversea=str(deli.get("fromOversea", "false")).lower() == "true",
                url=(it.get("url") or "").strip(),
                thumb=(it.get("thumb") or "").strip(),
            ))
        return items

    def search_all_pages(self, keyword: str, max_pages: int = 3, size: int = 50,
                          delay_sec: float = 0.5) -> List[DomeItem]:
        """여러 페이지를 순차적으로 모아서 반환 (과도한 호출 방지를 위해 delay 포함)"""
        all_items: List[DomeItem] = []
        for page in range(1, max_pages + 1):
            items = self.search_items(keyword, size=size, page=page)
            if not items:
                break
            all_items.extend(items)
            time.sleep(delay_sec)
        return all_items

    def cheapest(self, keyword: str, top_n: int = 5, max_pages: int = 2) -> List[DomeItem]:
        """키워드 검색 결과 중 배송비 포함 최저가 상위 N개 반환"""
        items = self.search_all_pages(keyword, max_pages=max_pages)
        items.sort(key=lambda x: x.price + x.deli_fee)
        return items[:top_n]


def group_similar_and_sort(items: List[DomeItem], similarity_threshold: float = 0.55) -> List[DomeItem]:
    """
    상품명이 비슷한 것끼리(=거의 같은 제품을 여러 판매자가 올린 경우) 묶어서,
    그 그룹 안에서는 원가+배송비가 낮은 순으로 정렬한다.
    서로 다른 상품군끼리는 뒤섞이지 않도록, 그룹의 등장 순서(원래 API 응답 순서)는 유지한다.
    """
    import difflib

    def norm(title: str) -> str:
        # 숫자/단위/공백을 지워서 비교하면 "500ml", "300ml" 같은 용량 차이에 덜 민감해짐
        import re
        t = re.sub(r"[0-9]+(ml|kg|g|cm|mm|oz|개입|개)?", "", title)
        return re.sub(r"\s+", "", t)

    groups: List[List[DomeItem]] = []
    group_keys: List[str] = []
    for it in items:
        key = norm(it.title)
        matched_idx = None
        for i, gk in enumerate(group_keys):
            if difflib.SequenceMatcher(None, key, gk).ratio() >= similarity_threshold:
                matched_idx = i
                break
        if matched_idx is None:
            groups.append([it])
            group_keys.append(key)
        else:
            groups[matched_idx].append(it)

    result: List[DomeItem] = []
    for g in groups:
        g.sort(key=lambda x: x.price + x.deli_fee)
        result.extend(g)
    return result


# ---------------------------------------------------------------------------
# 마진 계산
# ---------------------------------------------------------------------------

def calc_required_selling_price(cost: float, target_margin: float = DEFAULT_TARGET_MARGIN,
                                 coupang_fee_rate: float = DEFAULT_COUPANG_FEE_RATE) -> Optional[int]:
    """
    목표 마진율을 달성하기 위한 쿠팡 판매가 계산.
    마진율 정의: (판매가 - 원가 - 쿠팡수수료) / 판매가
    반환값이 None이면 그 마진율 자체가 수학적으로 불가능한 경우 (수수료율이 목표마진보다 큰 경우)
    """
    denom = 1 - coupang_fee_rate - target_margin
    if denom <= 0:
        return None
    return round(cost / denom)


def calc_margin_at_price(cost: float, selling_price: float,
                          coupang_fee_rate: float = DEFAULT_COUPANG_FEE_RATE) -> float:
    """특정 판매가로 팔았을 때 실제 마진율(%) 계산 (예: 경쟁상품 시장가 기준으로 팔면 마진이 몇 %인지)"""
    if selling_price <= 0:
        return 0.0
    fee = selling_price * coupang_fee_rate
    profit = selling_price - cost - fee
    return round(profit / selling_price * 100, 1)


# ---------------------------------------------------------------------------
# 출력 / 저장
# ---------------------------------------------------------------------------

def print_items_with_margin(items: List[DomeItem], target_margin: float = DEFAULT_TARGET_MARGIN,
                             coupang_fee_rate: float = DEFAULT_COUPANG_FEE_RATE):
    """
    검색 결과에 '목표 마진율 달성 판매가'를 같이 출력.
    도매꾹은 원가가 높아 40% 마진이 어려울 수 있으므로, 목표가 안 나오는 항목도
    걸러내지 않고 그대로 보여줘서 판단은 사용자가 하도록 함 (목표치는 조정 가능).
    """
    print(f"\n(쿠팡 수수료 {coupang_fee_rate*100:.1f}% 가정, 목표 마진율 {target_margin*100:.0f}% 기준)")
    print(f"총 {len(items)}건 검색됨")
    print("-" * 100)
    for it in items:
        cost = it.price + it.deli_fee
        required_price = calc_required_selling_price(cost, target_margin, coupang_fee_rate)
        price_str = f"{required_price:,}원" if required_price is not None else "계산불가(수수료가 목표마진보다 큼)"
        print(f"[{it.no}] {it.title[:32]:<32} 원가(배송포함) {cost:>7,}원 → "
              f"목표마진 판매가: {price_str}  ({it.seller_id})")
    print("-" * 100)


def export_to_excel_with_images(items: List[DomeItem], out_path: str = "도매꾹_검색결과.xlsx",
                                 img_width: int = 80, img_height: int = 80,
                                 target_margin: float = DEFAULT_TARGET_MARGIN,
                                 coupang_fee_rate: float = DEFAULT_COUPANG_FEE_RATE):
    """
    검색 결과를 엑셀로 저장하면서 각 상품의 썸네일 이미지 + 마진 계산 결과를 함께 삽입.
    - 쿠팡 런칭 전 '실제로 같은 제품인지' 눈으로 비교
    - '이 원가로 목표마진이 나오는지' 확인
    - L1(목표마진율), L2(쿠팡수수료율) 셀 값을 엑셀에서 직접 바꾸면 G열이 자동 재계산됨
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    FONT = "Arial"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "도매꾹 검색결과"

    headers = ["이미지", "상품번호", "상품명", "원가", "배송비", "원가합계",
               "목표마진 판매가", "판매자", "상품 보기"]
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border

    widths = [12, 12, 40, 10, 10, 12, 16, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 조정 가능한 기준값 (수식이 참조: L1=목표마진율, L2=쿠팡수수료율)
    ws.cell(row=1, column=11, value="목표 마진율")
    ws.cell(row=1, column=12, value=target_margin)
    ws.cell(row=2, column=11, value="쿠팡 수수료율")
    ws.cell(row=2, column=12, value=coupang_fee_rate)
    for r in (1, 2):
        ws.cell(row=r, column=11).font = Font(name=FONT, size=9)
        ws.cell(row=r, column=12).font = Font(name=FONT, size=9, color="0000FF")
        ws.cell(row=r, column=12).fill = PatternFill("solid", fgColor="DCE6F1")
    ws.cell(row=3, column=11,
            value="※ L1(목표마진율), L2(쿠팡수수료율) 값을 바꾸면 G열이 자동 재계산됩니다").font = Font(
        name=FONT, italic=True, size=8, color="808080")

    failed_images = []

    for row, it in enumerate(items, start=2):
        ws.row_dimensions[row].height = img_height * 0.75  # 대략 픽셀->포인트 환산

        if it.thumb:
            try:
                resp = requests.get(it.thumb, timeout=8)
                resp.raise_for_status()
                img_data = io.BytesIO(resp.content)
                xl_img = XLImage(img_data)
                xl_img.width, xl_img.height = img_width, img_height
                ws.add_image(xl_img, f"A{row}")
            except Exception as e:
                failed_images.append((it.no, str(e)))
                ws.cell(row=row, column=1, value="(이미지 실패)")

        ws.cell(row=row, column=2, value=it.no)
        ws.cell(row=row, column=3, value=it.title)
        ws.cell(row=row, column=4, value=it.price)
        ws.cell(row=row, column=5, value=it.deli_fee)
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")  # 원가합계
        ws.cell(row=row, column=7,
                value=f'=IF((1-$L$2-$L$1)<=0,"계산불가",ROUND(F{row}/(1-$L$2-$L$1),0))')  # 목표마진 판매가
        ws.cell(row=row, column=8, value=it.seller_id)

        link_cell = ws.cell(row=row, column=9, value="상품 보기")
        if it.url:
            link_cell.hyperlink = it.url
            link_cell.font = Font(name=FONT, size=9, color="0563C1", underline="single")
        else:
            link_cell.value = ""

        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            cell.font, cell.alignment, cell.border = Font(name=FONT, size=9), center, border
        link_cell.alignment, link_cell.border = center, border

    ws.freeze_panes = "A2"
    wb.save(out_path)

    if failed_images:
        print(f"\n⚠ 이미지 다운로드 실패 {len(failed_images)}건:")
        for no, err in failed_images:
            print(f"  - 상품번호 {no}: {err}")

    return out_path


if __name__ == "__main__":
    saved_key = load_saved_key()
    if saved_key:
        API_KEY = saved_key
        print(f"저장된 API_KEY({saved_key[:6]}...)를 자동으로 사용합니다.")
    else:
        API_KEY = input("도매꾹 API_KEY를 입력하세요: ").strip()
        if API_KEY:
            save = input("다음에도 자동으로 쓰도록 저장할까요? (y/n): ").strip().lower()
            if save == "y":
                save_key(API_KEY)
                print(f"저장 완료: {_KEY_FILE}")

    if not API_KEY:
        print("API_KEY가 입력되지 않았습니다.")
    else:
        client = DomeggookClient(api_key=API_KEY)
        keyword = input("검색할 키워드를 입력하세요: ").strip()

        margin_input = input(f"목표 마진율(%)을 입력하세요 (엔터 시 기본 {int(DEFAULT_TARGET_MARGIN*100)}%): ").strip()
        target_margin = DEFAULT_TARGET_MARGIN if not margin_input else float(margin_input) / 100

        try:
            results = client.cheapest(keyword, top_n=10, max_pages=2)
            print_items_with_margin(results, target_margin=target_margin)
            if results:
                print("\n이미지 포함 엑셀로 저장 중... (상품 수에 따라 몇 초~수십 초 걸릴 수 있어요)")
                try:
                    out_path = export_to_excel_with_images(results, target_margin=target_margin)
                    print(f"저장 완료: {out_path}")
                except ImportError:
                    print("Pillow 라이브러리가 필요합니다. 명령창에서 아래 명령어를 실행한 뒤 다시 시도해주세요:")
                    print("  pip install Pillow")
        except Exception as e:
            print(f"오류 발생: {e}")

    # 더블클릭으로 실행했을 때 창이 바로 닫히지 않도록 대기
    input("\n종료하려면 Enter를 누르세요...")
